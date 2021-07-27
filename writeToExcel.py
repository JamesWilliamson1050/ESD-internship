import openpyxl
from openpyxl.utils import get_column_letter


def write_module_name(moduleName):
    next_column = sheet1.max_column + 1
    print('A' +  str(next_column))

    # Need to get the rows and columns combined






if __name__ == '__main__':
    wb = openpyxl.load_workbook("Modules.xlsx")  #
    sheet1 = wb['Sheet1']
    write_module_name("Computer Programming")
